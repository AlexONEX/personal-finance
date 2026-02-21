import argparse
import io
import os
import sys
import warnings
from datetime import datetime

import openpyxl
import requests
import urllib3
from bs4 import BeautifulSoup
from dotenv import load_dotenv

from src.connectors.sheets import get_sheets_client

from fetch_historic import BACKFILL_FROM

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings("ignore", message="Unverified HTTPS request")

load_dotenv()

TODOS_REM_URL = (
    "https://www.bcra.gob.ar/todos-relevamiento-de-expectativas-de-mercado-rem/"
)
BASE_URL = "https://www.bcra.gob.ar"
SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID")
REM_SHEET_NAME = "REM"

MONTHS_MAP = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def get_publication_links(since_date: tuple[int, int] | None = None) -> list[dict]:
    """Get all publication links from the 'todos' page starting from since_date."""
    if since_date is None:
        since_date = (BACKFILL_FROM.year, BACKFILL_FROM.month)
    print(f"Fetching publication links from {TODOS_REM_URL}...")
    r = requests.get(TODOS_REM_URL, verify=False, timeout=30)
    soup = BeautifulSoup(r.text, "html.parser")

    table = soup.find("table")
    if not table:
        print("ERROR: No table found in 'todos' page.")
        return []

    links = []
    rows = table.find_all("tr")[1:]  # Skip header

    for row in rows:
        cols = row.find_all("td")
        if len(cols) < 3:
            continue

        link_tag = cols[0].find("a", href=True)
        if not link_tag:
            continue

        pub_url = link_tag["href"]
        if not pub_url.startswith("http"):
            pub_url = BASE_URL + pub_url

        period_text = cols[2].get_text(strip=True)  # e.g. "Enero 2026"
        try:
            m_text, y_text = period_text.split()
            period_date = (int(y_text), MONTHS_MAP[m_text.lower()])

            if since_date and period_date < since_date:
                continue

            links.append({"url": pub_url, "period": period_text, "date": period_date})
        except Exception:
            # print(f"  Warning: skipping row with period {period_text}: {e}")
            continue

    # Return links sorted by date (oldest first)
    return sorted(links, key=lambda x: x["date"])


def get_xlsx_from_publication(pub_url: str) -> str | None:
    """Find the .xlsx link in a publication page."""
    try:
        r = requests.get(pub_url, verify=False, timeout=30)
        soup = BeautifulSoup(r.text, "html.parser")
        links = soup.find_all("a", href=True)
        for link in links:
            text = link.get_text().lower()
            href = link["href"].lower()
            if ("tablas" in text or "tablas" in href) and href.endswith(".xlsx"):
                xlsx_url = link["href"]
                if not xlsx_url.startswith("http"):
                    xlsx_url = BASE_URL + xlsx_url
                return xlsx_url
    except Exception as e:
        print(f"  ERROR fetching publication {pub_url}: {e}")
    return None


def download_and_parse_rem(url: str) -> list[float]:
    """Download the REM xlsx and extract horizontal inflation projections."""
    try:
        r = requests.get(url, verify=False, timeout=30)
        f = io.BytesIO(r.content)
        wb = openpyxl.load_workbook(f, data_only=True)
        sheet = wb.worksheets[0]

        projections = []
        # Filas 7 a 13: M hasta M+6
        for row in range(7, 14):
            val = sheet.cell(row=row, column=4).value  # Mediana
            if val is not None:
                projections.append(float(val) / 100.0)
            else:
                projections.append(0.0)

        # Fila 14: Próximos 12 meses
        val_12m = sheet.cell(row=14, column=4).value
        if val_12m is not None:
            projections.append(float(val_12m) / 100.0)
        else:
            projections.append(0.0)

        return projections
    except Exception as e:
        print(f"    ERROR parsing xlsx {url}: {e}")
        return []


def update_rem_sheet(all_data: dict[str, list[float]]):
    """Update the Google Sheet 'REM' with a horizontal matrix of reports."""
    if not SPREADSHEET_ID:
        print("ERROR: SPREADSHEET_ID not set in .env", file=sys.stderr)
        return

    client = get_sheets_client()
    ss = client.open_by_key(SPREADSHEET_ID)

    try:
        ws = ss.worksheet(REM_SHEET_NAME)
    except:
        ws = ss.add_worksheet(title=REM_SHEET_NAME, rows=100, cols=10)

    ws.clear()

    # Headers
    headers = [
        "Mes Reporte",
        "Mes M",
        "Mes M+1",
        "Mes M+2",
        "Mes M+3",
        "Mes M+4",
        "Mes M+5",
        "Mes M+6",
        "Próx. 12m",
    ]

    # Sort by report month
    sorted_months = sorted(all_data.keys())
    payload = []
    for m in sorted_months:
        payload.append([m] + all_data[m])

    # Metadata
    metadata = [
        ["Última Actualización", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["", ""],
    ]

    ws.update("A1", metadata, value_input_option="USER_ENTERED")
    ws.update("A3", [headers], value_input_option="USER_ENTERED")

    if payload:
        ws.update("A4", payload, value_input_option="USER_ENTERED")

        num_rows = len(payload)
        end_row = 4 + num_rows - 1

        # Formateo
        ws.format(
            f"B4:I{end_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}}
        )
        ws.format(
            f"A4:A{end_row}", {"numberFormat": {"type": "DATE", "pattern": "mmm-yy"}}
        )

    ws.format(
        "A3:I3",
        {
            "textFormat": {
                "bold": True,
                "foregroundColor": {"red": 1, "green": 1, "blue": 1},
            },
            "backgroundColor": {"red": 0.2, "green": 0.3, "blue": 0.4},
        },
    )

    print(
        f"✓ Sheet {REM_SHEET_NAME!r} updated with {len(payload)} monthly reports (matrix)."
    )


def main():
    default_since = BACKFILL_FROM.strftime("%Y-%m")
    parser = argparse.ArgumentParser(
        description="Fetch historical REM projections from BCRA."
    )
    parser.add_argument(
        "--since",
        type=str,
        default=default_since,
        help=f"Start month (YYYY-MM). Default: {default_since}.",
    )
    args = parser.parse_args()

    try:
        y, m = map(int, args.since.split("-"))
        since_date = (y, m)
    except:
        print("Invalid format for --since. Use YYYY-MM.")
        sys.exit(1)

    pubs = get_publication_links(since_date)
    print(f"Found {len(pubs)} publications since {args.since}.")

    all_reports = {}
    for pub in pubs:
        print(f"Processing report for {pub['period']}...")
        xlsx_link = get_xlsx_from_publication(pub["url"])
        if xlsx_link:
            projections = download_and_parse_rem(xlsx_link)
            if projections:
                # Normalizamos la fecha del reporte (Mes M)
                report_key = f"{pub['date'][0]}-{pub['date'][1]:02d}-01"
                all_reports[report_key] = projections
                print(f"  Extracted {len(projections)} projection columns.")
            else:
                print(f"  Warning: no data extracted from {xlsx_link}")
        else:
            print(f"  Warning: no .xlsx found for {pub['period']}")

    if all_reports:
        update_rem_sheet(all_reports)
    else:
        print("No reports found to update.")


if __name__ == "__main__":
    main()
