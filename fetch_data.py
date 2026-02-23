"""fetch_data.py

Script unificado para la obtención de datos del Ingresos Tracker.
Obtiene CER (BCRA), CCL (Ambito/dolarapi), SPY (yfinance) y proyecciones REM (BCRA).
"""

import argparse
import io
import logging
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime

import openpyxl
import requests
import urllib3
import yfinance as yf
from bs4 import BeautifulSoup
from dotenv import load_dotenv

from src.connectors.sheets import get_sheets_client

# Suppress only InsecureRequestWarning for BCRA (they have cert issues)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

load_dotenv()

SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID")
SERVICE_ACCOUNT = "service_account.json"

HISTORIC_SHEET = "historic_data"
REM_SHEET = "REM"
FIRST_DATA_ROW = 4
BACKFILL_FROM = date(2022, 1, 1)

BCRA_API_URL = "https://api.bcra.gob.ar/estadisticas/v4.0/Monetarias/30"
AMBITO_CCL_URL = "https://mercados.ambito.com//dolarrava/cl/grafico/{desde}/{hasta}"
BCRA_TODOS_REM_URL = (
    "https://www.bcra.gob.ar/todos-relevamiento-de-expectativas-de-mercado-rem/"
)
BCRA_BASE_URL = "https://www.bcra.gob.ar"

MONTHS_MAP = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

MONTHS_MAP_SHORT = {
    "ene": 1,
    "feb": 2,
    "mar": 3,
    "abr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dic": 12,
}


def fetch_cer(since: date, until: date) -> dict[date, float]:
    results = {}
    offset, limit = 0, 3000

    while True:
        try:
            # WARNING: BCRA API has certificate issues (known Argentina Central Bank infrastructure issue)
            # Using verify=False ONLY for BCRA endpoints as a necessary exception
            resp = requests.get(
                BCRA_API_URL,
                params={
                    "desde": since.isoformat(),  # YYYY-MM-DD format
                    "hasta": until.isoformat(),
                    "limit": limit,
                    "offset": offset,
                },
                timeout=30,
                verify=False,
            )
            resp.raise_for_status()
            body = resp.json()

            detalle = []
            for variable in body.get("results", []):
                detalle.extend(variable.get("detalle", []))

            for record in detalle:
                if "fecha" not in record or "valor" not in record:
                    logger.warning(f"CER: skipping malformed record: {record}")
                    continue
                try:
                    d = datetime.strptime(record["fecha"], "%Y-%m-%d").date()
                    results[d] = float(record["valor"])
                except (ValueError, TypeError) as e:
                    logger.warning(
                        f"CER: invalid date or value in record {record}: {e}"
                    )
                    continue

            total = body.get("metadata", {}).get("resultset", {}).get("count", 0)
            offset += len(detalle)
            if offset >= total or not detalle:
                break

        except requests.exceptions.SSLError as e:
            logger.error(f"CER: SSL verification failed: {e}")
            raise
        except requests.exceptions.RequestException as e:
            logger.error(f"CER: request failed: {e}")
            break
        except (KeyError, ValueError) as e:
            logger.error(f"CER: invalid response format: {e}")
            break

    return results


def fetch_ccl_ambito(since: date, until: date) -> dict[date, float]:
    url = AMBITO_CCL_URL.format(desde=since.isoformat(), hasta=until.isoformat())
    out = {}

    try:
        # User-Agent needed to bypass Ambito's bot detection
        resp = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        data = resp.json()

        if not isinstance(data, list):
            logger.error("CCL Ambito: unexpected response format (not a list)")
            return out

        for row in data[1:]:
            if not isinstance(row, list) or len(row) < 2:
                logger.warning(f"CCL Ambito: skipping malformed row: {row}")
                continue
            try:
                d = datetime.strptime(str(row[0]).strip(), "%d/%m/%Y").date()
                out[d] = float(row[1])
            except (ValueError, TypeError, IndexError) as e:
                logger.warning(f"CCL Ambito: invalid row {row}: {e}")
                continue

    except requests.exceptions.RequestException as e:
        logger.warning(f"CCL Ambito: request failed: {e}")
    except (ValueError, KeyError) as e:
        logger.error(f"CCL Ambito: invalid response format: {e}")

    return out


def fetch_ccl_today() -> tuple[date, float] | None:
    try:
        resp = requests.get(
            "https://dolarapi.com/v1/dolares/contadoconliqui", timeout=10
        )
        resp.raise_for_status()
        body = resp.json()

        venta = body.get("venta")
        fecha = body.get("fechaActualizacion", "")

        if not venta or not fecha:
            logger.warning("CCL today: missing venta or fecha in response")
            return None

        try:
            d = datetime.fromisoformat(fecha.replace("Z", "+00:00")).date()
            return d, float(venta)
        except (ValueError, TypeError) as e:
            logger.warning(f"CCL today: invalid date or venta format: {e}")
            return None

    except requests.exceptions.RequestException as e:
        logger.warning(f"CCL today: request failed: {e}")
        return None
    except (KeyError, ValueError) as e:
        logger.warning(f"CCL today: invalid response format: {e}")
        return None


def fetch_spy(since: date, until: date) -> dict[date, float]:
    """
    Obtiene precios históricos de cierre de SPY usando yfinance.

    Returns:
        dict[date, float]: Mapa de fecha -> precio de cierre
    """
    out = {}

    try:
        # Descargar datos diarios de SPY
        ticker = yf.Ticker("SPY")
        df = ticker.history(start=since, end=until, interval="1d")

        if df.empty:
            logger.warning(f"SPY: no data returned for {since} to {until}")
            return out

        # Convertir a dict[date, float] usando el precio de cierre
        for idx, row in df.iterrows():
            d = idx.date() if hasattr(idx, "date") else idx
            out[d] = float(row["Close"])

        logger.info(f"SPY: fetched {len(out)} days from {since} to {until}")

    except Exception as e:
        logger.error(f"SPY: failed to fetch data: {e}")

    return out


def get_rem_publication_links(since_date: tuple[int, int]) -> list[dict]:
    try:
        # BCRA has certificate issues
        r = requests.get(BCRA_TODOS_REM_URL, timeout=30, verify=False)
        r.raise_for_status()
    except requests.exceptions.RequestException as e:
        logger.error(f"REM: failed to fetch publication list: {e}")
        return []

    soup = BeautifulSoup(r.text, "html.parser")
    table = soup.find("table")
    if not table:
        logger.warning("REM: no table found in publication page")
        return []

    links = []
    for row in table.find_all("tr")[1:]:
        cols = row.find_all("td")
        if len(cols) < 3:
            continue

        link_tag = cols[0].find("a", href=True)
        if not link_tag:
            continue

        period_text = cols[2].get_text(strip=True)
        try:
            m_text, y_text = period_text.split()
            period_date = (int(y_text), MONTHS_MAP[m_text.lower()])

            if period_date < since_date:
                continue

            pub_url = link_tag["href"]
            if not pub_url.startswith("http"):
                pub_url = BCRA_BASE_URL + pub_url

            links.append({"url": pub_url, "date": period_date, "period": period_text})

        except (ValueError, KeyError) as e:
            logger.warning(f"REM: failed to parse row {period_text}: {e}")
            continue

    return sorted(links, key=lambda x: x["date"])


def get_xlsx_from_publication(pub_url: str) -> str | None:
    try:
        # BCRA has certificate issues
        r = requests.get(pub_url, timeout=30, verify=False)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        for link in soup.find_all("a", href=True):
            text, href = link.get_text().lower(), link["href"].lower()
            if ("tablas" in text or "tablas" in href) and href.endswith(".xlsx"):
                xlsx_url = link["href"]
                if not xlsx_url.startswith("http"):
                    xlsx_url = BCRA_BASE_URL + xlsx_url

                # Fix for BCRA bug: redirect internal dev links to production
                if "desa.bcra.net" in xlsx_url:
                    xlsx_url = xlsx_url.replace(
                        "sitiopublico.desa.bcra.net", "www.bcra.gob.ar"
                    )
                return xlsx_url

    except requests.exceptions.RequestException as e:
        logger.warning(f"REM: failed to fetch publication page {pub_url}: {e}")
    except Exception as e:
        logger.warning(f"REM: failed to parse publication page {pub_url}: {e}")

    return None


def download_and_parse_rem(url: str) -> list[float]:
    try:
        # BCRA has certificate issues
        r = requests.get(url, timeout=30, verify=False)
        r.raise_for_status()

        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        sheet = wb.worksheets[0]

        projections = []
        for row in range(7, 14):  # M a M+6
            val = sheet.cell(row=row, column=4).value
            if val is not None:
                try:
                    projections.append(float(val) / 100.0)
                except (ValueError, TypeError) as e:
                    logger.warning(
                        f"REM: invalid projection value at row {row}: {val} ({e})"
                    )
                    projections.append(0.0)
            else:
                projections.append(0.0)

        val_12m = sheet.cell(row=14, column=4).value
        if val_12m is not None:
            try:
                projections.append(float(val_12m) / 100.0)
            except (ValueError, TypeError) as e:
                logger.warning(f"REM: invalid 12m projection: {val_12m} ({e})")
                projections.append(0.0)
        else:
            projections.append(0.0)

        return projections

    except requests.exceptions.RequestException as e:
        logger.error(f"REM: failed to download XLSX from {url}: {e}")
        return []
    except Exception as e:
        logger.error(f"REM: failed to parse XLSX from {url}: {e}")
        return []


def get_last_date_from_sheet() -> date:
    """Obtiene la última fecha registrada en historic_data para actualizar desde ahí."""
    try:
        client = get_sheets_client()
        ss = client.open_by_key(SPREADSHEET_ID)
        ws_h = ss.worksheet(HISTORIC_SHEET)
        existing_rows = ws_h.get_all_values()[FIRST_DATA_ROW - 1 :]

        dates = []
        for row in existing_rows:
            if len(row) >= 1 and row[0]:
                try:
                    d = datetime.strptime(row[0], "%d/%m/%Y").date()
                    dates.append(d)
                except ValueError:
                    continue

        if dates:
            return max(dates)

    except Exception as e:
        logger.error(f"Failed to get last date from sheet: {e}")

    return BACKFILL_FROM


def get_last_rem_date_from_sheet() -> tuple[int, int]:
    """
    Obtiene la última fecha (año, mes) de REM para actualizar solo datos posteriores.
    Returns: (year, month) tuple, defaults to (2022, 1) if sheet is empty.
    """
    try:
        client = get_sheets_client()
        ss = client.open_by_key(SPREADSHEET_ID)
        ws_r = ss.worksheet(REM_SHEET)
        existing_rows = ws_r.get_all_values()[
            3:
        ]  # Skip metadata and headers (rows 1-3)

        dates = []
        for row in existing_rows:
            if len(row) >= 1 and row[0]:
                date_str = row[0].strip()
                try:
                    # Try YYYY-MM-DD format first (e.g., "2020-01-01")
                    if "-" in date_str and len(date_str.split("-")[0]) == 4:
                        parts = date_str.split("-")
                        year = int(parts[0])
                        month = int(parts[1])
                        dates.append((year, month))
                    # Try mmm-YY format (e.g., "ene-20")
                    elif "-" in date_str and len(date_str.split("-")[0]) <= 3:
                        parts = date_str.split("-")
                        month_str = parts[0].lower()
                        year_short = parts[1]

                        if month_str in MONTHS_MAP_SHORT:
                            month = MONTHS_MAP_SHORT[month_str]
                            # Convert YY to YYYY (assume 20XX for now)
                            year = 2000 + int(year_short)
                            dates.append((year, month))
                except (ValueError, IndexError, KeyError) as e:
                    logger.warning(f"Failed to parse REM date '{date_str}': {e}")
                    continue

        if dates:
            return max(dates)

    except Exception as e:
        logger.error(f"Failed to get last REM date from sheet: {e}")

    return (BACKFILL_FROM.year, BACKFILL_FROM.month)


def update_sheets(cer_data, ccl_data, spy_data, rem_reports):
    client = get_sheets_client()
    ss = client.open_by_key(SPREADSHEET_ID)

    if cer_data or ccl_data or spy_data:
        ws_h = ss.worksheet(HISTORIC_SHEET)

        existing_rows = ws_h.get_all_values()[FIRST_DATA_ROW - 1 :]
        data_map = {}
        for row in existing_rows:
            if len(row) >= 1 and row[0]:
                data_map[row[0]] = [
                    row[1] if len(row) > 1 else "",
                    row[2] if len(row) > 2 else "",
                    row[3] if len(row) > 3 else "",
                ]

        for d, val in cer_data.items():
            fmt_d = d.strftime("%d/%m/%Y")
            if fmt_d not in data_map:
                data_map[fmt_d] = ["", "", ""]
            data_map[fmt_d][0] = val

        for d, val in ccl_data.items():
            fmt_d = d.strftime("%d/%m/%Y")
            if fmt_d not in data_map:
                data_map[fmt_d] = ["", "", ""]
            data_map[fmt_d][1] = val

        for d, val in spy_data.items():
            fmt_d = d.strftime("%d/%m/%Y")
            if fmt_d not in data_map:
                data_map[fmt_d] = ["", "", ""]
            data_map[fmt_d][2] = val

        sorted_dates = sorted(
            data_map.keys(), key=lambda x: datetime.strptime(x, "%d/%m/%Y")
        )
        payload = [
            [d, data_map[d][0], data_map[d][1], data_map[d][2]] for d in sorted_dates
        ]

        ws_h.update(
            range_name=f"A{FIRST_DATA_ROW}:D{FIRST_DATA_ROW + len(payload) - 1}",
            values=payload,
            value_input_option="USER_ENTERED",
        )

    if rem_reports:
        ws_r = ss.worksheet(REM_SHEET)

        # Update "Última Actualización" timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws_r.update(
            range_name="B1", values=[[timestamp]], value_input_option="USER_ENTERED"
        )

        # Get existing data to find last row
        existing_rem = ws_r.get_all_values()[3:]  # Skip rows 1-3 (metadata + headers)
        existing_months = {row[0] for row in existing_rem if len(row) >= 1 and row[0]}

        # Filter out reports that already exist
        new_reports = {
            month: projs
            for month, projs in rem_reports.items()
            if month not in existing_months
        }

        if new_reports:
            # Find the next empty row
            next_row = 4 + len(existing_rem)

            # Prepare payload for new reports only
            sorted_new_months = sorted(new_reports.keys())
            payload = [[m] + new_reports[m] for m in sorted_new_months]

            # Append new data
            ws_r.update(
                range_name=f"A{next_row}:I{next_row + len(payload) - 1}",
                values=payload,
                value_input_option="USER_ENTERED",
            )
            logger.info(f"Added {len(payload)} new REM reports")
        else:
            logger.info("No new REM reports to add")


def main():
    parser = argparse.ArgumentParser(
        description="Fetch all data for Ingresos Tracker. "
        "Por defecto, actualiza desde la última fecha registrada en el sheet."
    )
    parser.add_argument(
        "--since",
        type=str,
        default=None,
        help="Fecha inicio YYYY-MM-DD (opcional, por defecto usa última fecha del sheet)",
    )
    args = parser.parse_args()

    # Validación de input
    if args.since:
        try:
            since_dt = datetime.strptime(args.since, "%Y-%m-%d").date()
        except ValueError:
            logger.error(f"Invalid date format: {args.since}. Use YYYY-MM-DD")
            return

        # Validar rango de fechas razonable
        if since_dt > date.today():
            logger.error("Start date cannot be in the future")
            return
        if since_dt < date(2000, 1, 1):
            logger.error("Start date seems unreasonably old (before 2000)")
            return
    else:
        since_dt = get_last_date_from_sheet()

    until_dt = date.today()
    print(f"Updating dataset from {since_dt} to {until_dt}...")

    # Paralelizar fetch de datos
    with ThreadPoolExecutor(max_workers=5) as executor:
        # Lanzar CER, CCL Ambito, CCL today, SPY en paralelo
        future_cer = executor.submit(fetch_cer, since_dt, until_dt)
        future_ccl = executor.submit(fetch_ccl_ambito, since_dt, until_dt)
        future_today = executor.submit(fetch_ccl_today)
        future_spy = executor.submit(fetch_spy, since_dt, until_dt)

        # Esperar resultados
        cer = future_cer.result()
        ccl = future_ccl.result()
        today = future_today.result()
        spy = future_spy.result()

        if today and today[0] >= since_dt:
            ccl[today[0]] = today[1]

    # REM: paralelizar procesamiento de publicaciones
    # Use last REM date from sheet to avoid reprocessing old data
    last_rem_date = get_last_rem_date_from_sheet()
    logger.info(f"Last REM date in sheet: {last_rem_date[0]}-{last_rem_date[1]:02d}")

    rem_links = get_rem_publication_links(last_rem_date)
    rem_reports = {}

    def process_rem_publication(pub):
        xlsx = get_xlsx_from_publication(pub["url"])
        if xlsx:
            projs = download_and_parse_rem(xlsx)
            if projs:
                return (f"{pub['date'][0]}-{pub['date'][1]:02d}-01", projs)
        return None

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = [executor.submit(process_rem_publication, pub) for pub in rem_links]
        for future in as_completed(futures):
            result = future.result()
            if result:
                month_key, projs = result
                rem_reports[month_key] = projs

    update_sheets(cer, ccl, spy, rem_reports)
    print("Dataset updated successfully")


if __name__ == "__main__":
    main()
