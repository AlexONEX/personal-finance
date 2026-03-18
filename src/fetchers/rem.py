"""Fetcher para REM (Relevamiento de Expectativas de Mercado) del BCRA.

Realiza web scraping de publicaciones mensuales del BCRA y parsea archivos Excel
para extraer proyecciones de inflación.
"""

import io
import logging

import openpyxl
import requests
from bs4 import BeautifulSoup

from src.config import API_URLS, FETCH_CONFIG, MONTHS_MAP

logger = logging.getLogger(__name__)


class REMFetcher:
    """Obtiene proyecciones de inflación REM desde publicaciones del BCRA.

    Nota: No implementa DataSource porque devuelve un formato diferente
    (dict[str, list[float]] en lugar de dict[date, float]).
    """

    def fetch(self, since_date: tuple[int, int]) -> dict[str, list[float]]:
        """Obtiene reportes REM desde una fecha específica.

        Args:
            since_date: Tupla (año, mes) desde donde obtener datos

        Returns:
            Diccionario de "YYYY-MM-DD" -> lista de 8 proyecciones
            [M, M+1, M+2, M+3, M+4, M+5, M+6, 12m]
        """
        links = self._get_publication_links(since_date)
        reports = {}

        for pub in links:
            xlsx_url = self._get_xlsx_from_publication(pub["url"])
            if xlsx_url:
                projections = self._download_and_parse_excel(xlsx_url)
                if projections:
                    month_key = f"{pub['date'][0]}-{pub['date'][1]:02d}-01"
                    reports[month_key] = projections

        logger.info(f"REM: fetched {len(reports)} reports since {since_date}")
        return reports

    def _get_publication_links(
        self, since_date: tuple[int, int]
    ) -> list[dict[str, any]]:
        """Obtiene links de publicaciones REM desde la página del BCRA."""
        try:
            r = requests.get(
                API_URLS["bcra_rem_publications"],
                timeout=FETCH_CONFIG["timeout_seconds"],
                verify=False,  # BCRA has cert issues
            )
            r.raise_for_status()
        except requests.exceptions.RequestException as e:
            logger.error(f"REM: failed to fetch publication list: {e}")
            return []

        soup = BeautifulSoup(r.text, "html.parser")
        table = soup.find("table")
        if not table:
            logger.warning("REM: no table found in publication page")
            return []

        links = []
        for row in table.find_all("tr")[1:]:
            cols = row.find_all("td")
            if len(cols) < 3:
                continue

            link_tag = cols[0].find("a", href=True)
            if not link_tag:
                continue

            period_text = cols[2].get_text(strip=True)
            try:
                m_text, y_text = period_text.split()
                period_date = (int(y_text), MONTHS_MAP[m_text.lower()])

                if period_date < since_date:
                    continue

                pub_url = link_tag["href"]
                if not pub_url.startswith("http"):
                    pub_url = API_URLS["bcra_rem_base"] + pub_url

                links.append(
                    {"url": pub_url, "date": period_date, "period": period_text}
                )

            except (ValueError, KeyError) as e:
                logger.warning(f"REM: failed to parse row {period_text}: {e}")
                continue

        return sorted(links, key=lambda x: x["date"])

    def _get_xlsx_from_publication(self, pub_url: str) -> str | None:
        """Extrae URL del archivo Excel desde una página de publicación."""
        try:
            r = requests.get(
                pub_url, timeout=FETCH_CONFIG["timeout_seconds"], verify=False
            )
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")

            for link in soup.find_all("a", href=True):
                text, href = link.get_text().lower(), link["href"].lower()
                if ("tablas" in text or "tablas" in href) and href.endswith(".xlsx"):
                    xlsx_url = link["href"]
                    if not xlsx_url.startswith("http"):
                        xlsx_url = API_URLS["bcra_rem_base"] + xlsx_url

                    # Fix for BCRA bug: redirect internal dev links to production
                    if "desa.bcra.net" in xlsx_url:
                        xlsx_url = xlsx_url.replace(
                            "sitiopublico.desa.bcra.net", "www.bcra.gob.ar"
                        )
                    return xlsx_url

        except requests.exceptions.RequestException as e:
            logger.warning(f"REM: failed to fetch publication page {pub_url}: {e}")
        except Exception as e:
            logger.warning(f"REM: failed to parse publication page {pub_url}: {e}")

        return None

    def _download_and_parse_excel(self, url: str) -> list[float]:
        """Descarga y parsea archivo Excel de proyecciones REM."""
        try:
            r = requests.get(url, timeout=FETCH_CONFIG["timeout_seconds"], verify=False)
            r.raise_for_status()

            wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
            sheet = wb.worksheets[0]

            projections = []
            for row in range(7, 14):  # M a M+6
                val = sheet.cell(row=row, column=4).value
                if val is not None:
                    try:
                        projections.append(float(val) / 100.0)
                    except (ValueError, TypeError) as e:
                        logger.warning(
                            f"REM: invalid projection value at row {row}: {val} ({e})"
                        )
                        projections.append(0.0)
                else:
                    projections.append(0.0)

            val_12m = sheet.cell(row=14, column=4).value
            if val_12m is not None:
                try:
                    projections.append(float(val_12m) / 100.0)
                except (ValueError, TypeError) as e:
                    logger.warning(f"REM: invalid 12m projection: {val_12m} ({e})")
                    projections.append(0.0)
            else:
                projections.append(0.0)

            return projections

        except requests.exceptions.RequestException as e:
            logger.error(f"REM: failed to download XLSX from {url}: {e}")
            return []
        except Exception as e:
            logger.error(f"REM: failed to parse XLSX from {url}: {e}")
            return []
